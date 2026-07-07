#!/usr/bin/env python3
"""
Combined-folds deployment training for SICAPv2 CONCH v2 with a clean cosine schedule.

This entry point intentionally does not run cross-validation. It trains one deployment
checkpoint on the deduplicated union of Val1..Val4 Train.xlsx rows and leaves the
held-out partition/Test/Test.xlsx for post-training evaluation.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import random
from datetime import datetime
from pathlib import Path
from functools import partial

import numpy as np
import torch
from openpyxl import load_workbook
from torch.amp import GradScaler

import wandb

from paths import PARTITION_DIR, default_checkpoint_dir
from training_conchv2 import (
    CLASS_NAMES,
    CONCH_HF_CHECKPOINT,
    DEFAULT_CONFIG,
    DEFAULT_SEED,
    OUTPUT_DIR,
    GuidedLoss,
    ModelEMA,
    _PIXEL_FRAC_PARTITION,
    _trainable_model,
    build_cosine_warmup_scheduler,
    build_model,
    make_train_dataloader,
    parse_tta_scales,
    resolve_normalization,
    set_seed,
    train_one_epoch,
)


FOLDS = ("Val1", "Val2", "Val3", "Val4")


def _read_image_names_from_xlsx(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(x) if x is not None else "" for x in next(rows)]
    except StopIteration as e:
        raise ValueError(f"Empty Excel file: {path}") from e

    try:
        image_idx = header.index("image_name")
    except ValueError as e:
        raise KeyError(f"{path} does not contain an 'image_name' column. Header={header}") from e

    names = []
    for row in rows:
        if row is None or image_idx >= len(row):
            continue
        value = row[image_idx]
        if value:
            names.append(str(value))
    wb.close()
    return names


def _patient_id_from_image_name(image_name: str) -> str:
    return str(image_name).split("_Block", 1)[0]


def _dedupe_preserve_order(items: list[str]) -> list[str]:
    seen = set()
    out = []
    for item in items:
        if item in seen:
            continue
        seen.add(item)
        out.append(item)
    return out


def collect_combined_train_names(folds: tuple[str, ...] = FOLDS) -> tuple[list[str], dict]:
    all_names = []
    source_counts = {}
    for fold in folds:
        split_path = PARTITION_DIR / "Validation" / fold / "Train.xlsx"
        names = _read_image_names_from_xlsx(split_path)
        source_counts[f"{fold}/Train.xlsx"] = len(names)
        all_names.extend(names)

    unique_names = _dedupe_preserve_order(all_names)
    unique_patients = sorted({_patient_id_from_image_name(name) for name in unique_names})
    return unique_names, {
        "folds": list(folds),
        "source_counts": source_counts,
        "total_rows_before_dedup": len(all_names),
        "unique_tiles": len(unique_names),
        "duplicates_removed": len(all_names) - len(unique_names),
        "unique_patients": len(unique_patients),
    }


def _parse_class_weights(text: str) -> list[float]:
    try:
        weights = [float(x.strip()) for x in text.split(",") if x.strip()]
    except ValueError as e:
        raise argparse.ArgumentTypeError(f"invalid class weights: {e}") from e
    if len(weights) != len(CLASS_NAMES):
        raise argparse.ArgumentTypeError(f"expected {len(CLASS_NAMES)} weights, got {len(weights)}")
    if any(w <= 0 for w in weights):
        raise argparse.ArgumentTypeError("class weights must all be > 0")
    return weights


def train_combined_folds(config: dict, device: torch.device, use_wandb: bool = True, dry_run: bool = False):
    print(f"\n{'=' * 60}\n  COMBINED-FOLDS TRAINING (no validation set)\n{'=' * 60}")
    train_names, data_summary = collect_combined_train_names()
    print(
        f"  [Combined] rows={data_summary['total_rows_before_dedup']} -> "
        f"{data_summary['unique_tiles']} unique tiles | patients={data_summary['unique_patients']} | "
        f"duplicates_removed={data_summary['duplicates_removed']}"
    )
    for source_key, count in data_summary["source_counts"].items():
        print(f"    {source_key}: {count}")

    expected_tiles = int(config.get("expected_unique_tiles", 9959))
    expected_patients = int(config.get("expected_unique_patients", 124))
    if data_summary["unique_tiles"] != expected_tiles or data_summary["unique_patients"] != expected_patients:
        raise RuntimeError(
            "Combined-folds data mismatch: "
            f"expected {expected_tiles} tiles/{expected_patients} patients, got "
            f"{data_summary['unique_tiles']} tiles/{data_summary['unique_patients']} patients."
        )

    train_loader = make_train_dataloader(train_names, config, label="combined_train")

    model = build_model(config).to(device)
    ema = None
    if config.get("use_ema", False):
        ema = ModelEMA(model, decay=float(config.get("ema_decay", 0.999)))
        print(f"  [EMA] enabled (decay={config.get('ema_decay', 0.999)})")

    use_compile = config.get("use_compile")
    if use_compile is None:
        import platform
        use_compile = platform.system() != "Windows"
    if use_compile and int(torch.__version__.split(".")[0]) >= 2:
        try:
            import platform
            backend = "inductor" if platform.system() != "Windows" else "aot_eager"
            model = torch.compile(model, backend=backend)
            print("  [Compile] torch.compile enabled.")
        except Exception as e:
            print(f"  [Compile] unavailable: {e}")
    elif not use_compile:
        print("  [Compile] disabled.")

    trainable_model = _trainable_model(model)
    ema_source = trainable_model if ema is not None else None
    criterion = GuidedLoss(config["class_weights"], config["dice_weight"], config["ce_weight"]).to(device)

    encoder_params = [p for p in trainable_model.encoder.parameters() if p.requires_grad]
    decoder_params = list(trainable_model.decoder.parameters())
    param_groups = []
    if encoder_params:
        param_groups.append({"params": encoder_params, "lr": config["learning_rate"] / 10, "label": "encoder"})
    param_groups.append({"params": decoder_params, "lr": config["learning_rate"], "label": "decoder"})
    optimizer = torch.optim.AdamW(param_groups, weight_decay=config["weight_decay"])

    ga = int(config.get("grad_accum_steps", 1))
    steps_per_epoch = max(1, math.ceil(len(train_loader) / max(1, ga)))
    max_epochs = 1 if dry_run else int(config["max_epochs"])
    total_optimizer_steps = steps_per_epoch * max_epochs
    scheduler = build_cosine_warmup_scheduler(
        optimizer,
        total_steps=total_optimizer_steps,
        warmup_pct=float(config.get("warmup_pct", 0.07)),
        min_lr_ratio=float(config.get("cosine_min_lr_ratio", 0.01)),
    )
    print(
        f"  [Scheduler] cosine_warmup over {total_optimizer_steps} optimizer steps "
        f"({steps_per_epoch} steps/epoch x {max_epochs} epochs)"
    )

    scaler = GradScaler("cuda", enabled=(device.type == "cuda"))
    out_dir = Path(config.get("output_dir", OUTPUT_DIR))
    out_dir.mkdir(parents=True, exist_ok=True)

    history = []
    for epoch in range(1, max_epochs + 1):
        lr_dec = optimizer.param_groups[-1]["lr"]
        print(f"\n  Epoch {epoch}/{max_epochs}  (decoder lr={lr_dec:.2e}, accum={ga})")
        train_loss = train_one_epoch(
            model,
            train_loader,
            criterion,
            optimizer,
            scaler,
            device,
            grad_accum_steps=ga,
            scheduler=scheduler,
            step_scheduler_per_batch=True,
            ema=ema,
            ema_source=ema_source,
        )
        lr_encoder = optimizer.param_groups[0]["lr"]
        lr_decoder = optimizer.param_groups[-1]["lr"]
        history.append({
            "epoch": epoch,
            "train_loss": float(train_loss),
            "lr_encoder": float(lr_encoder),
            "lr_decoder": float(lr_decoder),
        })
        print(f"  Train Loss: {train_loss:.4f} | lr_decoder={lr_decoder:.2e}")

        if use_wandb and wandb.run is not None:
            wandb.log({
                "combined/train_loss": train_loss,
                "combined/lr_encoder": lr_encoder,
                "combined/lr_decoder": lr_decoder,
                "epoch": epoch,
            })

    output_name = str(config.get("output_name", "final_all_folds_cosine15.pth"))
    if not output_name.endswith(".pth"):
        output_name += ".pth"
    ckpt_path = out_dir / output_name
    to_save = ema.ema.state_dict() if ema is not None else _trainable_model(model).state_dict()
    torch.save(to_save, ckpt_path)

    metadata = {
        "checkpoint_path": str(ckpt_path),
        "checkpoint_source": "EMA" if ema is not None else "raw",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "data_summary": data_summary,
        "history": history,
        "config": {
            "batch_size": config["batch_size"],
            "grad_accum_steps": config.get("grad_accum_steps", 1),
            "effective_batch": config["batch_size"] * config.get("grad_accum_steps", 1),
            "learning_rate": config["learning_rate"],
            "weight_decay": config["weight_decay"],
            "max_epochs": max_epochs,
            "scheduler": "cosine_warmup",
            "warmup_pct": config.get("warmup_pct", 0.07),
            "cosine_min_lr_ratio": config.get("cosine_min_lr_ratio", 0.01),
            "class_weights": list(config["class_weights"]),
            "dice_weight": config["dice_weight"],
            "ce_weight": config["ce_weight"],
            "unfreeze_last": config["unfreeze_last"],
            "use_weighted_sampler": config.get("use_weighted_sampler", False),
            "sampler_replacement": config.get("sampler_replacement", False),
            "norm_mean": list(config["norm_mean"]),
            "norm_std": list(config["norm_std"]),
            "color_aug_enabled": config.get("color_aug_enabled", True),
            "use_ema": config.get("use_ema", False),
            "ema_decay": config.get("ema_decay", 0.999),
            "seed": config.get("seed", DEFAULT_SEED),
        },
    }
    metadata_path = ckpt_path.with_suffix(".json")
    with metadata_path.open("w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)

    print(f"\n  [OK] Final checkpoint saved: {ckpt_path}")
    print(f"  [OK] Metadata saved:         {metadata_path}")
    return ckpt_path, metadata_path, metadata


def main() -> None:
    parser = argparse.ArgumentParser(description="Train combined-folds CONCH v2 deployment model with cosine warmup.")
    parser.add_argument("--max-epochs", type=int, default=15)
    parser.add_argument("--unfreeze-last", type=int, default=4)
    parser.add_argument("--learning-rate", type=float, default=4e-5)
    parser.add_argument("--class-weights", type=_parse_class_weights, default=_parse_class_weights("1.0,3.586,2.573,4.21"))
    parser.add_argument("--ema", action="store_true", default=True)
    parser.add_argument("--no-ema", action="store_false", dest="ema")
    parser.add_argument("--ema-decay", type=float, default=0.999)
    parser.add_argument("--batch-size", type=int, default=6)
    parser.add_argument("--grad-accum", type=int, default=2)
    parser.add_argument("--warmup-pct", type=float, default=0.07)
    parser.add_argument("--cosine-min-lr-ratio", type=float, default=0.01)
    parser.add_argument("--weight-decay", type=float, default=None)
    parser.add_argument("--num-workers", type=int, default=None)
    parser.add_argument("--no-weighted-sampler", action="store_true")
    parser.add_argument("--sampler-replacement", action="store_true")
    parser.add_argument("--no-color-aug", action="store_true")
    parser.add_argument("--imagenet-norm", action="store_true")
    parser.add_argument("--weights", type=str, default=None, help="Local CONCH backbone checkpoint.")
    parser.add_argument("--hf-token", type=str, default=None)
    parser.add_argument("--compile", action="store_true")
    parser.add_argument("--no-compile", action="store_true")
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--output-name", type=str, default="final_all_folds_cosine15.pth")
    parser.add_argument("--wandb-project", type=str, default="SICAPv2_CONCH_maskLUT")
    parser.add_argument("--wandb-name", type=str, default="combined_cosine15_runI_config")
    parser.add_argument("--no-wandb", action="store_true")
    parser.add_argument("--dry-run", action="store_true", help="Run one epoch and save a dry-run checkpoint.")
    parser.add_argument("--data-summary-only", action="store_true")
    parser.add_argument("--tta-scales", type=str, default="1.0", help="Parsed only for config parity; no validation is run.")
    args = parser.parse_args()

    set_seed(args.seed)
    random.seed(args.seed)
    np.random.seed(args.seed)

    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.kernel32.SetThreadExecutionState(0x80000000 | 0x00000001)
            print("[System] Anti-sleep enabled.")
        except Exception:
            pass

    train_names, data_summary = collect_combined_train_names()
    if args.data_summary_only:
        print(json.dumps(data_summary, indent=2))
        return
    del train_names

    config = DEFAULT_CONFIG.copy()
    config["seed"] = args.seed
    config["max_epochs"] = int(args.max_epochs)
    config["unfreeze_last"] = int(args.unfreeze_last)
    config["learning_rate"] = float(args.learning_rate)
    config["class_weights"] = args.class_weights
    config["use_ema"] = bool(args.ema)
    config["ema_decay"] = float(args.ema_decay)
    config["batch_size"] = int(args.batch_size)
    config["grad_accum_steps"] = max(1, int(args.grad_accum))
    config["warmup_pct"] = float(args.warmup_pct)
    config["cosine_min_lr_ratio"] = float(args.cosine_min_lr_ratio)
    config["use_weighted_sampler"] = not args.no_weighted_sampler
    config["sampler_replacement"] = bool(args.sampler_replacement)
    config["color_aug_enabled"] = not args.no_color_aug
    config["conch_checkpoint"] = args.weights
    config["conch_hf_token"] = args.hf_token
    config["output_dir"] = args.output_dir.resolve() if args.output_dir else default_checkpoint_dir("checkpoints_conch_masklut")
    config["output_name"] = args.output_name
    config["expected_unique_tiles"] = 9959
    config["expected_unique_patients"] = 124
    config["tta_scales"] = parse_tta_scales(args.tta_scales)
    config["use_cosine_schedule"] = True
    if args.weight_decay is not None:
        config["weight_decay"] = float(args.weight_decay)
    if args.num_workers is not None:
        config["num_workers"] = int(args.num_workers)
    if args.compile and args.no_compile:
        print("[WARN] --compile and --no-compile provided; using --no-compile.")
    if args.no_compile:
        config["use_compile"] = False
    elif args.compile:
        config["use_compile"] = True

    norm_mean, norm_std, norm_warning = resolve_normalization(args.imagenet_norm)
    config["use_imagenet_norm"] = args.imagenet_norm
    config["norm_mean"] = norm_mean
    config["norm_std"] = norm_std
    if norm_warning:
        print(f"[WARN] {norm_warning}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")
    print(f"CONCH backbone source: {args.weights or CONCH_HF_CHECKPOINT}")
    print(
        f"Combined cosine train: epochs={config['max_epochs']} | bs={config['batch_size']} | "
        f"accum={config['grad_accum_steps']} | lr={config['learning_rate']} | "
        f"unfreeze_last={config['unfreeze_last']} | ema={config['use_ema']}"
    )
    print(
        f"  class_weights={config['class_weights']} | scheduler=cosine_warmup | "
        f"warmup_pct={config['warmup_pct']} | min_lr_ratio={config['cosine_min_lr_ratio']}"
    )
    print(f"  output={config['output_dir'] / config['output_name']}")

    use_wandb = not args.no_wandb
    if use_wandb and not args.dry_run:
        try:
            wandb.init(
                project=args.wandb_project,
                name=args.wandb_name,
                config={
                    "script": "training_combined_cosine",
                    "folds": list(FOLDS),
                    "data_summary": data_summary,
                    "checkpoint_name": config["output_name"],
                    "conch_checkpoint": args.weights or CONCH_HF_CHECKPOINT,
                    "pixel_frac_partition": [float(x) for x in _PIXEL_FRAC_PARTITION],
                    **{k: v for k, v in config.items() if isinstance(v, (int, float, str, bool, list, tuple))},
                },
                tags=["combined_folds", "cosine15", "deployment", "CONCH", "SICAPv2"],
            )
        except Exception as e:
            print(f"[WARN] Weights & Biases unavailable: {e}")
            use_wandb = False

    ckpt_path, metadata_path, metadata = train_combined_folds(
        config,
        device,
        use_wandb=use_wandb,
        dry_run=args.dry_run,
    )
    if use_wandb and wandb.run is not None:
        wandb.run.summary["checkpoint_path"] = str(ckpt_path)
        wandb.run.summary["metadata_path"] = str(metadata_path)
        if metadata["history"]:
            wandb.run.summary["final_train_loss"] = metadata["history"][-1]["train_loss"]
            wandb.run.summary["final_lr_decoder"] = metadata["history"][-1]["lr_decoder"]
        wandb.finish()


if __name__ == "__main__":
    main()
