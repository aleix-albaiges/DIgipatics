#!/usr/bin/env python3
"""
Evaluate a trained CONCH v2 multiclass segmentation checkpoint on partition/Test.

Default target:
  artifacts/checkpoints_conch_masklut/final_all_folds.pth
  partition/Test/Test.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import types
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "src"
VENV_SITE_PACKAGES = ROOT / "prostata_env" / "Lib" / "site-packages"


def bootstrap_runtime() -> None:
    os.environ.setdefault("KMP_DUPLICATE_LIB_OK", "TRUE")
    for module_name in ("numpy", "pandas", "scipy", "PIL"):
        try:
            __import__(module_name)
        except Exception:
            pass
    if VENV_SITE_PACKAGES.exists():
        sys.path.insert(0, str(VENV_SITE_PACKAGES))
    if SRC.exists():
        sys.path.insert(0, str(SRC))
    sys.modules.setdefault("wandb", types.ModuleType("wandb"))


bootstrap_runtime()

import numpy as np
import torch
from openpyxl import load_workbook

import sicap_imports  # noqa: F401
from paths import PARTITION_DIR, default_checkpoint_dir
from training_conchv2 import (
    CLASS_NAMES,
    CONCH_HF_CHECKPOINT,
    DEFAULT_CONFIG,
    GuidedLoss,
    NUM_CLASSES,
    build_model,
    make_val_dataloader,
    print_aggregated_matrices,
    resolve_normalization,
    validate_one_epoch,
)


def _read_image_names_from_xlsx(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = [str(x) if x is not None else "" for x in next(rows)]
    except StopIteration as e:
        raise ValueError(f"Empty Excel file: {path}") from e

    try:
        image_idx = header.index("image_name")
    except ValueError as e:
        raise KeyError(f"{path} does not contain an 'image_name' column. Header={header}") from e

    names = []
    for row in rows:
        if row is None or image_idx >= len(row):
            continue
        value = row[image_idx]
        if value:
            names.append(str(value))
    wb.close()
    return names


def _safe_torch_load(path: Path, device: torch.device):
    try:
        return torch.load(path, map_location=device, weights_only=True)
    except TypeError:
        return torch.load(path, map_location=device)


def _strip_compile_prefix(state_dict: dict) -> dict:
    if not any(k.startswith("_orig_mod.") for k in state_dict):
        return state_dict
    return {k.replace("_orig_mod.", "", 1): v for k, v in state_dict.items()}


def _metrics_from_cm(cm: np.ndarray) -> dict:
    per_class = {}
    f1_values = []
    for i, name in enumerate(CLASS_NAMES):
        tp = int(cm[i, i])
        fp = int(cm[:, i].sum() - tp)
        fn = int(cm[i, :].sum() - tp)
        precision = tp / (tp + fp + 1e-8)
        recall = tp / (tp + fn + 1e-8)
        f1 = 2.0 * precision * recall / (precision + recall + 1e-8)
        per_class[name] = {
            "precision": float(precision),
            "recall": float(recall),
            "f1": float(f1),
            "support_pixels": int(cm[i, :].sum()),
        }
        f1_values.append(f1)

    nc_nc = int(cm[0, 0])
    nc_c = int(cm[0, 1:].sum())
    c_nc = int(cm[1:, 0].sum())
    c_c = int(cm[1:, 1:].sum())
    cancer_precision = c_c / (c_c + nc_c + 1e-8)
    cancer_recall = c_c / (c_c + c_nc + 1e-8)
    cancer_f1 = 2.0 * cancer_precision * cancer_recall / (cancer_precision + cancer_recall + 1e-8)
    accuracy = (c_c + nc_nc) / (c_c + nc_nc + nc_c + c_nc + 1e-8)

    return {
        "macro_f1_4class": float(np.mean(f1_values)),
        "per_class": per_class,
        "binary_cancer_vs_nc": {
            "precision": float(cancer_precision),
            "recall": float(cancer_recall),
            "f1": float(cancer_f1),
            "accuracy": float(accuracy),
            "confusion_matrix": {
                "tn_nc_as_nc": nc_nc,
                "fp_nc_as_cancer": nc_c,
                "fn_cancer_as_nc": c_nc,
                "tp_cancer_as_cancer": c_c,
            },
        },
    }


def _write_per_class_csv(path: Path, metrics: dict) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["class", "precision", "recall", "f1", "support_pixels"])
        writer.writeheader()
        for class_name, values in metrics["per_class"].items():
            writer.writerow({"class": class_name, **values})


def _load_metadata_config(checkpoint_path: Path) -> dict:
    metadata_path = checkpoint_path.with_suffix(".json")
    if not metadata_path.exists():
        return {}
    with metadata_path.open("r", encoding="utf-8") as f:
        metadata = json.load(f)
    return metadata.get("config", {})


def main() -> None:
    parser = argparse.ArgumentParser(description="Evaluate CONCH v2 checkpoint on partition/Test.")
    parser.add_argument(
        "--checkpoint",
        type=Path,
        default=default_checkpoint_dir("checkpoints_conch_masklut") / "final_all_folds.pth",
        help="Model .pth checkpoint to evaluate.",
    )
    parser.add_argument(
        "--test-xlsx",
        type=Path,
        default=PARTITION_DIR / "Test" / "Test.xlsx",
        help="Excel split file containing image_name rows.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="Directory for metrics JSON/CSV. Default: checkpoint_dir/test_eval.",
    )
    parser.add_argument("--batch-size", type=int, default=None, help="Override eval batch size.")
    parser.add_argument("--num-workers", type=int, default=0, help="DataLoader workers for evaluation.")
    parser.add_argument("--weights", type=str, default=None, help="Local CONCH backbone checkpoint, if needed.")
    parser.add_argument("--hf-token", type=str, default=None, help="HF token for CONCH checkpoint access.")
    parser.add_argument("--imagenet-norm", action="store_true", help="Use ImageNet normalization instead of CONCH/CLIP.")
    parser.add_argument("--tta", action="store_true", help="Use D4 TTA during test evaluation.")
    parser.add_argument("--tta-scales", type=str, default="1.0", help='Comma-separated scales, e.g. "0.875,1.0,1.125".')
    args = parser.parse_args()

    checkpoint_path = args.checkpoint.resolve()
    test_xlsx = args.test_xlsx.resolve()
    if not checkpoint_path.exists():
        raise FileNotFoundError(f"Checkpoint not found: {checkpoint_path}")
    if not test_xlsx.exists():
        raise FileNotFoundError(f"Test split not found: {test_xlsx}")

    metadata_config = _load_metadata_config(checkpoint_path)
    config = DEFAULT_CONFIG.copy()
    config.update({k: v for k, v in metadata_config.items() if k in config})
    config["num_workers"] = int(args.num_workers)
    config["use_weighted_sampler"] = False
    config["conch_checkpoint"] = args.weights
    config["conch_hf_token"] = args.hf_token
    if args.batch_size is not None:
        config["batch_size"] = int(args.batch_size)

    norm_mean, norm_std, norm_warning = resolve_normalization(args.imagenet_norm)
    if not args.imagenet_norm and metadata_config.get("norm_mean") and metadata_config.get("norm_std"):
        norm_mean = metadata_config["norm_mean"]
        norm_std = metadata_config["norm_std"]
    config["norm_mean"] = norm_mean
    config["norm_std"] = norm_std

    names = _read_image_names_from_xlsx(test_xlsx)
    if not names:
        raise ValueError(f"No image_name rows found in {test_xlsx}")

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"Device: {device}")
    print(f"Checkpoint: {checkpoint_path}")
    print(f"Test split: {test_xlsx} ({len(names)} images)")
    print(f"CONCH backbone source: {args.weights or CONCH_HF_CHECKPOINT}")
    if norm_warning:
        print(f"[WARN] {norm_warning}")

    loader = make_val_dataloader(names, config)
    model = build_model(config)
    state_dict = _strip_compile_prefix(_safe_torch_load(checkpoint_path, device=torch.device("cpu")))
    missing, unexpected = model.load_state_dict(state_dict, strict=False)
    if missing or unexpected:
        print(f"[WARN] load_state_dict strict=False | missing={len(missing)} | unexpected={len(unexpected)}")
        if missing:
            print(f"  missing examples: {missing[:5]}")
        if unexpected:
            print(f"  unexpected examples: {unexpected[:5]}")
    model.to(device)

    criterion = GuidedLoss(config["class_weights"], config["dice_weight"], config["ce_weight"]).to(device)
    tta_config = None
    if args.tta:
        scales = tuple(float(x.strip()) for x in args.tta_scales.split(",") if x.strip())
        tta_config = {"use_d4": True, "scales": scales}
        print(f"TTA enabled: D4=True scales={scales}")

    val_loss, val_metrics = validate_one_epoch(model, loader, criterion, device, tta_config=tta_config)
    cm = val_metrics["confusion_matrix"]
    metrics = _metrics_from_cm(cm)

    print(f"\nTest loss: {val_loss:.4f}")
    print(f"4-class macro F-score: {metrics['macro_f1_4class']:.4f}")
    for class_name, values in metrics["per_class"].items():
        print(
            f"  {class_name}: F1={values['f1']:.4f} "
            f"P={values['precision']:.4f} R={values['recall']:.4f} "
            f"support={values['support_pixels']}"
        )
    binary = metrics["binary_cancer_vs_nc"]
    print(
        f"Binary cancer F-score: {binary['f1']:.4f} "
        f"P={binary['precision']:.4f} R={binary['recall']:.4f} Acc={binary['accuracy']:.4f}"
    )
    print_aggregated_matrices(cm)

    output_dir = (args.output_dir or (checkpoint_path.parent / "test_eval")).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)
    result = {
        "checkpoint": str(checkpoint_path),
        "test_xlsx": str(test_xlsx),
        "num_images": len(names),
        "loss": float(val_loss),
        "tta": bool(args.tta),
        "tta_scales": list(tta_config["scales"]) if tta_config else [1.0],
        "confusion_matrix_4class": cm.tolist(),
        "metrics": metrics,
        "config": {
            "batch_size": config["batch_size"],
            "num_workers": config["num_workers"],
            "unfreeze_last": config["unfreeze_last"],
            "class_weights": list(config["class_weights"]),
            "dice_weight": config["dice_weight"],
            "ce_weight": config["ce_weight"],
            "norm_mean": list(config["norm_mean"]),
            "norm_std": list(config["norm_std"]),
        },
    }
    json_path = output_dir / "test_metrics.json"
    csv_path = output_dir / "test_per_class_f1.csv"
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(result, f, indent=2)
    _write_per_class_csv(csv_path, metrics)
    print(f"\nSaved metrics: {json_path}")
    print(f"Saved per-class CSV: {csv_path}")


if __name__ == "__main__":
    main()
