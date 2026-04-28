"""
Class distribution:

- **Patch (Excel)**: columns NC, G3, G4, G5 en los .xlsx de ``partition/`` → índice
  0..3 vía ``sicap_mapping.excel_row_to_class_index`` (same logic as analysis / no G4C en el índice).

- **Pixel (mask)**: gris → ``build_mask_lut(gg5_gray_min)`` + optional
  ``clean_all_speckles`` (same logic as ``training_conch``).

By default only includes ``image_name`` listed in the partition (Train/Test .xlsx).

Usage:
  python scripts/analyze_masks.py
  python scripts/analyze_masks.py --all-masks
  python analyze_masks.py --gg5-gray-min 170 --mask-clean-min-area 16
"""
from __future__ import annotations

import argparse
from collections import Counter
from pathlib import Path

import sicap_imports  # noqa: F401

import cv2
import numpy as np
import openpyxl

from sicap_mapping import (
    CLASS_NAMES as NAMES,
    build_mask_lut,
    clean_all_speckles,
    excel_row_to_class_index,
)

try:
    from sicap_mapping import DEFAULT_GG5_GRAY_MIN
except ImportError:
    DEFAULT_GG5_GRAY_MIN = 170

from paths import MASKS_DIR, PARTITION_DIR

# Same default as training_conch DEFAULT_CONFIG
_DEFAULT_CLEAN_AREA = 16


def _print_class_abs_pct(
    title: str,
    counts: dict[int, int] | Counter,
    total: int,
    *,
    mutually_exclusive: bool = True,
) -> None:
    """Tabla: clase | id | absolutos | % del total (denominador = total)."""
    print(title)
    if not mutually_exclusive:
        print("  (Clases no excluyentes: un mismo parche puede contar en varias rows; los % no suman 100%.)\n")
    print(f"  {'clase':<6} {'id':>4}  {'absolutos':>14}  {'%':>8}")
    print("  " + "-" * 44)
    s = 0
    for c in range(4):
        n = int(counts.get(c, 0))
        s += n
        pct = 100.0 * n / total if total else 0.0
        print(f"  {NAMES[c]:<6} {c:>4}  {n:>14,}  {pct:8.2f}")
    print("  " + "-" * 44)
    if mutually_exclusive:
        print(f"  {'SUMA':<6} {'':>4}  {s:>14,}  {100.0 * s / total if total else 0.0:8.2f}")
    else:
        print(f"  {'SUMA':<6} {'':>4}  {s:>14,}  {'---':>8}")
    print(f"  Resumen absolutos: {_one_line_abs(dict(counts))}")


def _one_line_abs(counts: dict[int, int]) -> str:
    return ", ".join(f"{NAMES[c]}={counts.get(c, 0):,}" for c in range(4))


def partition_xlsx_files() -> list[Path]:
    patterns = [
        "Validation/*/Train.xlsx",
        "Validation/*/Test.xlsx",
        "Test/Train.xlsx",
        "Test/Test.xlsx",
    ]
    files: list[Path] = []
    for pat in patterns:
        files.extend(sorted(PARTITION_DIR.glob(pat)))
    return files


def read_partition_rows(path: Path) -> tuple[dict[str, int], list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    rows = list(it)
    wb.close()
    return idx, rows


def row_to_excel_class(row: tuple, idx: dict[str, int]) -> int | None:
    return excel_row_to_class_index(
        row[idx["NC"]],
        row[idx["G3"]],
        row[idx["G4"]],
        row[idx["G5"]],
    )


def _read_image_names_xlsx(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return []
        header = [str(h).strip() if h is not None else "" for h in header]
        if "image_name" not in header:
            return []
        i = header.index("image_name")
        out: list[str] = []
        for row in rows:
            if row is None or i >= len(row):
                continue
            v = row[i]
            if v is None:
                continue
            s = str(v).strip()
            if s:
                out.append(s)
        return out
    finally:
        wb.close()


def collect_partition_image_names() -> set[str]:
    names: set[str] = set()
    for xp in partition_xlsx_files():
        for n in _read_image_names_xlsx(xp):
            names.add(n)
    return names


def excel_class_distributions() -> tuple[Counter, Counter, int, int]:
    """
    Devuelve:
    - conteo por clase incluyendo **cada row** Excel (puede repetir image_name entre folds),
    - conteo por clase con **image_name únicos** (última aparición al recorrer files ordenados gana),
    - nº rows con clase None / ambigua,
    - nº image_name únicos.
    """
    per_row: Counter = Counter()
    unique_last: dict[str, int] = {}
    n_none_row = 0

    for xp in partition_xlsx_files():
        idx, rows = read_partition_rows(xp)
        for row in rows:
            if "image_name" not in idx:
                continue
            name = str(row[idx["image_name"]]).strip()
            if not name:
                continue
            ec = row_to_excel_class(row, idx)
            if ec is None:
                n_none_row += 1
                continue
            per_row[ec] += 1
            unique_last[name] = ec

    counter_unique = Counter(unique_last.values())
    return per_row, counter_unique, n_none_row, len(unique_last)


def run(
    all_masks: bool,
    gg5_gray_min: int,
    mask_clean_min_area: int,
) -> None:
    lut = build_mask_lut(gg5_gray_min)

    print("=== MAPEO (como training_conch) ===")
    print(f"  build_mask_lut(gg5_gray_min={gg5_gray_min})")
    print(f"  clean_all_speckles min_area={mask_clean_min_area} (0 = disabledo)")
    print()

    print("=== PARCHES - etiqueta Excel (NC, G3, G4, G5 -> indice 0..3) ===")
    per_row, uniq, n_none, n_unique = excel_class_distributions()
    total_rows = sum(per_row.values()) + n_none
    print(f"Filas Excel en partition/ (total): {total_rows:,} | sin clase (None): {n_none:,}")
    if n_none:
        print(f"  (% y SUMA usan denominador {total_rows:,} rows; las {n_none:,} None no tienen row de clase.)")
    print()
    _print_class_abs_pct(
        "Por row Excel (cada aparicion en un .xlsx; puede repetir image_name entre folds):",
        per_row,
        total_rows,
    )
    print()

    print(f"image_name unicos en partition/: {n_unique:,}")
    total_u = sum(uniq.values())
    print()
    _print_class_abs_pct(
        "Por parche unico (una etiqueta por image_name; ultima aparicion en Excel gana):",
        uniq,
        total_u,
    )
    print()

    if all_masks:
        all_paths = sorted(MASKS_DIR.glob("*.jpg")) + sorted(MASKS_DIR.glob("*.png"))
        label = "all masks in masks/"
    else:
        allowed = collect_partition_image_names()
        if not allowed:
            raise SystemExit(
                f"No se encontraron image_name en {PARTITION_DIR}. "
                "¿Falta la carpeta partition o los .xlsx?"
            )
        all_paths = []
        missing_in_masks: list[str] = []
        for name in sorted(allowed):
            p = MASKS_DIR / name
            if p.is_file():
                all_paths.append(p)
            else:
                missing_in_masks.append(name)
        label = f"only patches en partición ({len(allowed)} unique names en .xlsx)"
        print(f"Nombres únicos en partición (Excel): {len(allowed):,}")
        print(f"Files de mask encontrados: {len(all_paths):,}")
        if missing_in_masks:
            print(
                f"Warning: {len(missing_in_masks)} nombres en partición sin file en masks/ "
                f"(ej.: {missing_in_masks[:3]})"
            )

    total_pixels = {0: 0, 1: 0, 2: 0, 3: 0}
    patch_has_class = {0: 0, 1: 0, 2: 0, 3: 0}
    majority_per_patch: Counter[int] = Counter()

    n_ok = 0
    n_fail = 0
    for mp in all_paths:
        buf = np.fromfile(str(mp), dtype=np.uint8)
        mask = cv2.imdecode(buf, cv2.IMREAD_GRAYSCALE)
        if mask is None:
            n_fail += 1
            continue
        n_ok += 1
        m = lut[mask].astype(np.int64)
        if mask_clean_min_area > 0:
            m = clean_all_speckles(m, min_area=mask_clean_min_area)
        flat = m.ravel()
        bc = np.bincount(flat.astype(np.int64), minlength=4)
        for c in range(4):
            total_pixels[c] += int(bc[c])
            if bc[c] > 0:
                patch_has_class[c] += 1
        majority_per_patch[int(np.argmax(bc))] += 1

    grand_total = sum(total_pixels.values())
    print("\n=== PIXELES - mascaras + LUT + limpieza ===")
    print(f"Modo: {label}")
    print(f"Mascaras leidas OK: {n_ok:,} | fallidas: {n_fail}")
    print(f"Total pixeles (suma clases): {grand_total:,}\n")

    _print_class_abs_pct(
        "Pixeles por clase (absolutos = recuento de pixeles; % sobre total pixeles):",
        {c: total_pixels[c] for c in range(4)},
        grand_total,
    )
    print()

    ph = {c: patch_has_class[c] for c in range(4)}
    _print_class_abs_pct(
        f"Patches con >=1 pixel de cada clase (absolutos = n. de mascaras; % sobre {n_ok:,} mascaras leidas):",
        ph,
        n_ok,
        mutually_exclusive=False,
    )
    print()

    maj = {c: int(majority_per_patch.get(c, 0)) for c in range(4)}
    _print_class_abs_pct(
        "Patches segun clase mayoritaria en la mascara (absolutos = n. de mascaras):",
        maj,
        n_ok,
    )


def main() -> None:
    ap = argparse.ArgumentParser(description="Excel distribution (patch) and masks (pixel)")
    ap.add_argument(
        "--all-masks",
        action="store_true",
        help="Usar todos los files en masks/ (ignorar partition)",
    )
    ap.add_argument(
        "--gg5-gray-min",
        type=int,
        default=None,
        help=f"Umbral gris para GG5 (default: {DEFAULT_GG5_GRAY_MIN})",
    )
    ap.add_argument(
        "--mask-clean-min-area",
        type=int,
        default=None,
        help=f"Minimum component area para cualquier clase (default: {_DEFAULT_CLEAN_AREA}; 0 = without cleaning)",
    )
    args = ap.parse_args()

    g5 = DEFAULT_GG5_GRAY_MIN if args.gg5_gray_min is None else args.gg5_gray_min
    clean_a = (
        _DEFAULT_CLEAN_AREA
        if args.mask_clean_min_area is None
        else int(args.mask_clean_min_area)
    )

    run(all_masks=args.all_masks, gg5_gray_min=g5, mask_clean_min_area=clean_a)


if __name__ == "__main__":
    main()
