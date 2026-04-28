"""
Checks consistency between:
  - wsi_labels.xlsx (Gleason per slide)
  - Excel de partition/ (NC, G3, G4, G5 + G4C: G4 subtype; see sicap_mapping)
  - Masks in masks/ with sicap_mapping.MASK_LUT

Run: python scripts/audit_partition_vs_masks.py
"""
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path

import cv2
import numpy as np
import openpyxl

import sicap_imports  # noqa: F401

from paths import MASKS_DIR, PARTITION_DIR, get_project_root
from sicap_mapping import CLASS_NAMES, MASK_LUT as LUT, excel_row_to_class_index

WSI_LABELS = get_project_root() / "wsi_labels.xlsx"

NAMES = list(CLASS_NAMES)


def gleason_to_wsi_group(gp: int, gs: int) -> str:
    """Grade-group style cluster used previously (patch-level coarse)."""
    if (gp, gs) in [(4, 5), (5, 4), (5, 5)]:
        return "GG5"
    if (gp, gs) in [(4, 4), (3, 5), (5, 3)]:
        return "GG4"
    if (gp, gs) in [(3, 4), (4, 3)]:
        return "GG3"
    return "NC/other"


def load_wsi_labels() -> dict[str, tuple[int, int, str]]:
    """slide_id -> (gp, gs, group_str)"""
    wb = openpyxl.load_workbook(WSI_LABELS, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it)
    h = [str(x).strip() if x is not None else "" for x in header]
    i_s = h.index("slide_id")
    i_p = h.index("Gleason_primary")
    i_s2 = h.index("Gleason_secondary")
    out: dict[str, tuple[int, int, str]] = {}
    for row in it:
        if row[i_s] is None:
            continue
        sid = str(row[i_s]).strip()
        gp = int(row[i_p])
        gs = int(row[i_s2])
        out[sid] = (gp, gs, gleason_to_wsi_group(gp, gs))
    wb.close()
    return out


def partition_xlsx_files() -> list[Path]:
    patterns = [
        "Validation/*/Train.xlsx",
        "Validation/*/Test.xlsx",
        "Test/Train.xlsx",
        "Test/Test.xlsx",
    ]
    files: list[Path] = []
    for pat in patterns:
        files.extend(sorted(PARTITION_DIR.glob(pat)))
    return files


def read_partition_rows(path: Path) -> tuple[dict[str, int], list[tuple]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = next(it)
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
    rows = list(it)
    wb.close()
    return idx, rows


def row_to_excel_class(row: tuple, idx: dict[str, int]) -> int | None:
    """Delegates to sicap_mapping.excel_row_to_class_index (G4C does not affect the index)."""
    return excel_row_to_class_index(
        row[idx["NC"]],
        row[idx["G3"]],
        row[idx["G4"]],
        row[idx["G5"]],
    )


def mask_majority_class(mask_path: Path) -> tuple[int | None, np.ndarray]:
    buf = np.fromfile(str(mask_path), dtype=np.uint8)
    m = cv2.imdecode(buf, cv2.IMREAD_GRAYSCALE)
    if m is None:
        return None, np.zeros(4, dtype=np.int64)
    mapped = LUT[m]
    bc = np.bincount(mapped.ravel().astype(np.int64), minlength=4)
    return int(np.argmax(bc)), bc


def main() -> None:
    wsi = load_wsi_labels()
    xfiles = partition_xlsx_files()
    print(f"wsi_labels.xlsx: {len(wsi)} slides")
    print(f"Files partition .xlsx: {len(xfiles)}")

    # Una entrada por image_name (misma etiqueta en todos los .xlsx)
    all_rows: list[tuple[str, tuple, dict[str, int]]] = []
    for xp in xfiles:
        idx, rows = read_partition_rows(xp)
        for row in rows:
            name = str(row[idx["image_name"]]).strip()
            if not name:
                continue
            all_rows.append((name, row, idx))

    print(f"Filas totales en todos los Excel: {len(all_rows):,}")

    unique_items: dict[str, tuple[int | None, tuple, dict[str, int]]] = {}
    conflicts = 0
    for name, row, idx in all_rows:
        ec = row_to_excel_class(row, idx)
        if name not in unique_items:
            unique_items[name] = (ec, row, idx)
        else:
            ec0, _, _ = unique_items[name]
            if ec != ec0:
                conflicts += 1
    print(f"image_name unicos: {len(unique_items):,}")
    print(f"Conflictos (mismo image_name, distinta clase Excel): {conflicts}")

    agree_maj = 0
    agree_presence = 0
    missing_mask = 0
    bad_mask = 0
    cm = np.zeros((4, 4), dtype=np.int64)  # excel_class x mask_majority

    # wsi vs modo de excel por slide
    slide_excel_classes: dict[str, list[int]] = defaultdict(list)

    for name, (ec, row, idx) in unique_items.items():
        if ec is None:
            continue
        slide = name.split("_")[0]
        slide_excel_classes[slide].append(ec)

        mp = MASKS_DIR / name
        if not mp.is_file():
            missing_mask += 1
            continue
        maj, _ = mask_majority_class(mp)
        if maj is None:
            bad_mask += 1
            continue
        cm[ec, maj] += 1
        if maj == ec:
            agree_maj += 1
        # Presence: tumor si hay >=1 pixel de esa clase; NC si no hay pixel G3/G4/G5
        buf = np.fromfile(str(mp), dtype=np.uint8)
        m2 = cv2.imdecode(buf, cv2.IMREAD_GRAYSCALE)
        mapped2 = LUT[m2]
        if ec == 0:
            if not np.any((mapped2 == 1) | (mapped2 == 2) | (mapped2 == 3)):
                agree_presence += 1
        else:
            if np.any(mapped2 == ec):
                agree_presence += 1

    n = int(cm.sum())
    print("\n=== Máscara (LUT) vs clase derivada del Excel ===")
    print(f"Patches con Excel válido y mask OK: {n:,}")
    print(f"Coinciden (mayoría mask == Excel): {agree_maj:,} ({100.0 * agree_maj / n if n else 0:.2f}%)")
    print(
        f"Coinciden (presencia): mask tiene >=1 pixel de la clase Excel (NC=tumor sin pixel G3/G4/G5): "
        f"{agree_presence:,} ({100.0 * agree_presence / n if n else 0:.2f}%)"
    )
    print(f"Missing file mask: {missing_mask:,} | decode failed: {bad_mask:,}")
    print("\nConfusion matrix [row=Excel, column=Máscara mayoritaria]")
    print("       " + "  ".join(f"{NAMES[j]:>6s}" for j in range(4)))
    for i in range(4):
        print(f"{NAMES[i]:>6s} " + "  ".join(f"{cm[i,j]:>6d}" for j in range(4)))

    # Relation wsi_labels (slide) vs distribución Excel en patches
    print("\n=== wsi_labels (grupo slide) vs etiqueta Excel en patches ===")
    group_order = ["GG5", "GG4", "GG3", "NC/other"]
    for g in group_order:
        slides = [s for s, (_, _, gg) in wsi.items() if gg == g]
        if not slides:
            continue
        c = Counter()
        tot = 0
        for s in slides:
            for cl in slide_excel_classes.get(s, []):
                c[cl] += 1
                tot += 1
        print(f"  {g} ({len(slides)} slides, {tot:,} patches en partition con esos slides):")
        if tot == 0:
            print("    (no patch de esos slides en partition)")
            continue
        for cl in range(4):
            print(f"    {NAMES[cl]}: {c[cl]:,} ({100.0 * c[cl] / tot:.1f}%)")

    # Slides en partition pero no en wsi_labels
    part_slides = set(s for s in slide_excel_classes)
    wsi_slides = set(wsi.keys())
    only_part = part_slides - wsi_slides
    only_wsi = wsi_slides - part_slides
    if only_part:
        print(f"\nSlides con patches en partition pero sin row en wsi_labels: {len(only_part)} (ej. {list(only_part)[:5]})")
    if only_wsi:
        print(f"Slides en wsi_labels sin no patch en partition: {len(only_wsi)}")

    ex_g5_not_wsi_gg5 = 0
    ex_cancer_on_nc_slide = 0
    for name, (ec, _, _) in unique_items.items():
        if ec is None:
            continue
        slide = name.split("_")[0]
        if slide not in wsi:
            continue
        _, _, gg = wsi[slide]
        if ec == 3 and gg != "GG5":
            ex_g5_not_wsi_gg5 += 1
        if ec in (1, 2, 3) and gg == "NC/other":
            ex_cancer_on_nc_slide += 1
    print("\n=== Cruce wsi_labels (grupo slide) vs etiqueta Excel del patch ===")
    print(f"Patches Excel=G5 pero slide (wsi) no es GG5: {ex_g5_not_wsi_gg5:,}")
    print(f"Patches Excel con tumor (G3/G4/G5) pero slide clasificada NC/other en wsi: {ex_cancer_on_nc_slide:,}")


if __name__ == "__main__":
    main()
